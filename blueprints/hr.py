"""
hr.py -- 인건비/연차 관리 Blueprint.
직원 관리, 급여 관리, 연차 관리, 급여 항목/보험 요율 관리.
직원/급여는 admin만, 연차는 admin+manager.
"""
import os
import tempfile

from flask import (
    Blueprint, render_template, request, current_app,
    jsonify, send_file,
)
from flask_login import login_required, current_user
from datetime import date, timedelta

from auth import role_required, _log_action
from db_utils import get_db

hr_bp = Blueprint('hr', __name__, url_prefix='/hr')


def _auto_recalc_payroll(db, employee_id):
    """급여 항목 변경 후 해당 직원의 draft 급여가 있으면 자동 재계산.
    현재 월(또는 가장 최근 draft)의 급여를 재계산한다.
    """
    try:
        today = date.today()
        pay_month = f'{today.year}-{today.month:02d}'
        existing = db.query_payroll(pay_month=pay_month)
        payroll_rec = next(
            (r for r in existing
             if r.get('employee_id') == int(employee_id)
             and r.get('status') == 'draft'),
            None
        )
        if payroll_rec:
            db.recalculate_payroll(payroll_rec['id'])
            return True
    except Exception as e:
        print(f"[HR] auto recalc payroll error (emp={employee_id}): {e}")
    return False


# ══════════════════════════════════════════════
#  직원 관리
# ══════════════════════════════════════════════

@hr_bp.route('/employees')
@role_required('admin', 'general')
def employees():
    """직원 관리 메인 페이지"""
    return render_template('hr/employees.html')


@hr_bp.route('/api/employees')
@role_required('admin', 'general')
def api_employees():
    """직원 목록 JSON API"""
    db = get_db()
    status = request.args.get('status', '')
    try:
        rows = db.query_employees(status=status or None)
        # 각 직원에 법정 연차일수 추가
        for r in rows:
            r['legal_leave_days'] = db.calculate_legal_leave_days(
                r.get('hire_date')
            )
        return jsonify({'success': True, 'employees': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/employees', methods=['POST'])
@role_required('admin', 'general')
def api_create_employee():
    """직원 등록"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    name = (data.get('name') or '').strip()
    hire_date = (data.get('hire_date') or '').strip()

    if not name or not hire_date:
        return jsonify({'error': '이름과 입사일은 필수입니다.'}), 400

    try:
        base_salary = float(data.get('base_salary', 0))
    except (ValueError, TypeError):
        return jsonify({'error': '기본급이 올바르지 않습니다.'}), 400

    try:
        dependents_count = int(data.get('dependents_count', 1))
    except (ValueError, TypeError):
        dependents_count = 1

    payload = {
        'name': name,
        'position': (data.get('position') or '').strip(),
        'department': (data.get('department') or '').strip(),
        'base_salary': base_salary,
        'hire_date': hire_date,
        'status': (data.get('status') or '재직').strip(),
        'memo': (data.get('memo') or '').strip(),
        'bank_name': (data.get('bank_name') or '').strip(),
        'bank_account': (data.get('bank_account') or '').strip(),
        'dependents_count': dependents_count,
        'is_tax_exempt': bool(data.get('is_tax_exempt', False)),
    }

    try:
        result = db.insert_employee(payload)
        _log_action('create_employee', target=name,
                     detail=f'입사일={hire_date}, 기본급={base_salary:,.0f}',
                     new_value=payload)
        return jsonify({'success': True, 'employee': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/employees/<int:emp_id>', methods=['PUT'])
@role_required('admin', 'general')
def api_update_employee(emp_id):
    """직원 수정"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    name = (data.get('name') or '').strip()
    hire_date = (data.get('hire_date') or '').strip()

    if not name or not hire_date:
        return jsonify({'error': '이름과 입사일은 필수입니다.'}), 400

    try:
        base_salary = float(data.get('base_salary', 0))
    except (ValueError, TypeError):
        return jsonify({'error': '기본급이 올바르지 않습니다.'}), 400

    try:
        dependents_count = int(data.get('dependents_count', 1))
    except (ValueError, TypeError):
        dependents_count = 1

    payload = {
        'name': name,
        'position': (data.get('position') or '').strip(),
        'department': (data.get('department') or '').strip(),
        'base_salary': base_salary,
        'hire_date': hire_date,
        'status': (data.get('status') or '재직').strip(),
        'memo': (data.get('memo') or '').strip(),
        'bank_name': (data.get('bank_name') or '').strip(),
        'bank_account': (data.get('bank_account') or '').strip(),
        'dependents_count': dependents_count,
        'is_tax_exempt': bool(data.get('is_tax_exempt', False)),
    }

    try:
        result = db.update_employee(emp_id, payload)
        _log_action('update_employee', target=f'{name} (id={emp_id})',
                     detail=f'입사일={hire_date}, 기본급={base_salary:,.0f}',
                     new_value=payload)
        return jsonify({'success': True, 'employee': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/employees/<int:emp_id>/retire', methods=['POST'])
@role_required('admin', 'general')
def api_retire_employee(emp_id):
    """직원 퇴사 처리 (삭제하지 않고 상태만 변경)"""
    db = get_db()
    data = request.get_json() or {}
    retire_date = (data.get('retire_date') or '').strip()
    memo = (data.get('memo') or '').strip()

    try:
        update_data = {'status': '퇴사'}
        if retire_date:
            update_data['retire_date'] = retire_date
        if memo:
            update_data['memo'] = memo
        db.update_employee(emp_id, update_data)
        _log_action('retire_employee', target=f'id={emp_id}',
                     detail=f'퇴사일={retire_date}', new_value=update_data)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  직원 엑셀 일괄 등록
# ══════════════════════════════════════════════

# ═══ 엑셀 헤더 매핑 ═══
# [기본정보] 컬럼: (한글헤더, DB컬럼, 필수여부)
_EMP_BASIC_COLS = [
    ('이름',      'name',             True),
    ('부서',      'department',       False),
    ('직급',      'position',         False),
    ('입사일',    'hire_date',        True),
    ('기본급',    'base_salary',      False),
    ('부양가족수', 'dependents_count', False),
    ('급여은행',   'bank_name',       False),
    ('계좌번호',   'bank_account',    False),
    ('상태',      'status',           False),
    ('비과세',    'is_tax_exempt',    False),
    ('메모',      'memo',             False),
]

# [급여항목] 컬럼: (한글헤더, component_type, is_taxable)
# is_fixed=True (매월 고정), amount는 엑셀에서 입력된 값
_EMP_COMPONENT_COLS = [
    ('직급수당', 'position_allowance',      True),
    ('연차수당', 'annual_leave_allowance',  True),
    ('연장수당', 'overtime_pay',            True),
    ('기타수당', 'other_allowance',         True),
    ('식대',    'meal_allowance',           False),
    ('차량보조', 'vehicle_allowance',        False),
    ('공제금액', 'deduction',                False),
]

# 전체 컬럼 (템플릿 생성/업로드 공통)
_EMP_EXCEL_COLS = (
    [(h, d, req) for h, d, req in _EMP_BASIC_COLS] +
    [(h, c, False) for h, c, _ in _EMP_COMPONENT_COLS]
)


@hr_bp.route('/employees/template')
@role_required('admin', 'general')
def employees_template():
    """직원 일괄등록용 엑셀 템플릿 다운로드."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치'}), 500

    wb = Workbook()
    ws = wb.active
    ws.title = '직원명단'

    # 2단 헤더: 1행=그룹구분 / 2행=컬럼명
    n_basic = len(_EMP_BASIC_COLS)
    n_comp = len(_EMP_COMPONENT_COLS)
    group_row = ['[기본정보]'] + [''] * (n_basic - 1) + ['[급여항목]'] + [''] * (n_comp - 1)
    headers = [c[0] for c in _EMP_BASIC_COLS] + [c[0] for c in _EMP_COMPONENT_COLS]

    ws.append(group_row)
    ws.append(headers)

    # 그룹 셀 병합
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_basic)
    ws.merge_cells(start_row=1, start_column=n_basic + 1,
                   end_row=1, end_column=n_basic + n_comp)

    # 헤더 스타일
    basic_fill = PatternFill('solid', start_color='D9E8F5')
    comp_fill = PatternFill('solid', start_color='FFF3CD')
    for col_idx in range(1, n_basic + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True, size=11)
        ws.cell(row=1, column=col_idx).fill = basic_fill
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col_idx).font = Font(bold=True)
        ws.cell(row=2, column=col_idx).fill = basic_fill
        ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center')
    for col_idx in range(n_basic + 1, n_basic + n_comp + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True, size=11)
        ws.cell(row=1, column=col_idx).fill = comp_fill
        ws.cell(row=1, column=col_idx).alignment = Alignment(horizontal='center')
        ws.cell(row=2, column=col_idx).font = Font(bold=True)
        ws.cell(row=2, column=col_idx).fill = comp_fill
        ws.cell(row=2, column=col_idx).alignment = Alignment(horizontal='center')

    # 예시 2줄
    # 기본정보: 이름, 부서, 직급, 입사일, 기본급, 부양가족수, 급여은행, 계좌번호, 상태, 비과세, 메모
    # 급여항목: 직급수당, 연차수당, 연장수당, 기타수당, 식대, 차량보조, 공제금액
    ws.append(['홍길동', '운영팀', '대리', '2025-03-01', 3000000, 2,
               '국민은행', '123-456-789012', '재직', 'N', '샘플입니다',
               100000, 50000, 0, 0, 200000, 100000, 0])
    ws.append(['김영희', '생산팀', '사원', '2025-06-15', 2500000, 1,
               '신한은행', '111-222-333333', '재직', 'Y', '',
               0, 0, 0, 0, 200000, 0, 0])

    # 가이드 시트
    ws2 = wb.create_sheet('작성가이드')
    guide = [
        ['구분', '컬럼명', '필수', '설명'],
        ['기본정보', '이름', 'O', '직원 이름 (중복 시 입사일로 구분)'],
        ['기본정보', '부서', 'X', '운영팀/생산팀 등'],
        ['기본정보', '직급', 'X', '대리/사원/과장 등'],
        ['기본정보', '입사일', 'O', 'YYYY-MM-DD 형식 (예: 2025-04-21)'],
        ['기본정보', '기본급', 'X', '숫자만 입력 (원단위, 콤마 없이)'],
        ['기본정보', '부양가족수', 'X', '본인 포함 숫자 (빈값=1)'],
        ['기본정보', '급여은행', 'X', '급여 이체 은행명'],
        ['기본정보', '계좌번호', 'X', '급여 이체 계좌번호'],
        ['기본정보', '상태', 'X', '재직/퇴사, 빈값=재직'],
        ['기본정보', '비과세', 'X', 'Y=비과세 대상, N=과세 (빈값=N)'],
        ['기본정보', '메모', 'X', '비고사항'],
        ['급여항목', '직급수당', 'X', '매월 고정 금액 (원단위, 과세)'],
        ['급여항목', '연차수당', 'X', '매월 고정 금액 (원단위, 과세)'],
        ['급여항목', '연장수당', 'X', '매월 고정 금액 (원단위, 과세)'],
        ['급여항목', '기타수당', 'X', '매월 고정 금액 (원단위, 과세)'],
        ['급여항목', '식대', 'X', '매월 고정 금액 (원단위, 비과세 월 20만원 한도)'],
        ['급여항목', '차량보조', 'X', '매월 고정 금액 (원단위, 비과세)'],
        ['급여항목', '공제금액', 'X', '매월 고정 공제 (원단위)'],
        ['', '', '', ''],
        ['', '※ 샘플 2줄 삭제 후 실제 데이터 입력하세요.', '', ''],
        ['', '※ 1행은 그룹 구분(병합), 2행은 실제 컬럼명 → 삭제 금지', '', ''],
        ['', '※ 이름과 입사일은 반드시 입력해야 합니다.', '', ''],
        ['', '※ 동일 이름 + 입사일이 이미 등록되어 있으면 스킵됩니다.', '', ''],
        ['', '※ 급여항목 미입력/0 → 해당 항목 미등록 (기존 있으면 0원 덮어씀)', '', ''],
    ]
    for row in guide:
        ws2.append(row)
    for col_idx in range(1, 5):
        ws2.cell(row=1, column=col_idx).font = Font(bold=True)
        ws2.cell(row=1, column=col_idx).fill = basic_fill

    # 열 너비 — MergedCell(병합셀) 회피: index 기반으로 column_letter 구함
    from openpyxl.utils import get_column_letter
    for ws_ in (ws, ws2):
        max_col = ws_.max_column or 1
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws_[letter]:
                # 병합셀 skip
                if cell.__class__.__name__ == 'MergedCell':
                    continue
                v = cell.value
                if v is None:
                    continue
                L = len(str(v))
                if L > max_len:
                    max_len = L
            ws_.column_dimensions[letter].width = min(max_len + 2, 40)

    # 임시파일 저장 후 전송
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()
    return send_file(tmp.name, as_attachment=True,
                     download_name='직원등록_템플릿.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@hr_bp.route('/api/employees/bulk-upload', methods=['POST'])
@role_required('admin', 'general')
def api_bulk_upload_employees():
    """직원 엑셀 일괄 등록."""
    if 'file' not in request.files:
        return jsonify({'error': '파일이 없습니다.'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'error': '파일명이 없습니다.'}), 400
    if not f.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Excel 파일(.xlsx, .xls)만 업로드 가능합니다.'}), 400

    try:
        from openpyxl import load_workbook
    except ImportError:
        return jsonify({'error': 'openpyxl 미설치'}), 500

    db = get_db()

    # 기존 직원 조회 (중복 체크용)
    try:
        existing = db.query_employees() or []
    except Exception:
        existing = []
    existing_keys = set()
    for e in existing:
        key = (str(e.get('name', '')).strip(),
               str(e.get('hire_date', ''))[:10])
        if key[0] and key[1]:
            existing_keys.add(key)

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'error': f'엑셀 파일 읽기 실패: {e}'}), 400

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    # 헤더 행 탐색: '이름'이 포함된 첫 행을 헤더로 인식 (1행=그룹, 2행=컬럼명 형식 지원)
    header_row_idx = None
    for idx, r in enumerate(rows[:3]):  # 최대 3행까지 탐색
        vals = [str(v or '').strip() for v in r]
        if '이름' in vals and '입사일' in vals:
            header_row_idx = idx
            break
    if header_row_idx is None:
        return jsonify({
            'error': "헤더 행을 찾을 수 없습니다. '이름' + '입사일' 컬럼이 필요합니다."
        }), 400

    header = [str(h or '').strip() for h in rows[header_row_idx]]
    col_map = {}
    for idx, h in enumerate(header):
        if h:
            col_map[h] = idx

    # 필수 헤더 확인
    missing_headers = [c[0] for c in _EMP_BASIC_COLS
                       if c[2] and c[0] not in col_map]
    if missing_headers:
        return jsonify({
            'error': f'필수 헤더 누락: {", ".join(missing_headers)}'
        }), 400

    # 데이터 시작 인덱스 = header_row_idx + 1
    data_start = header_row_idx + 1

    def _get(row, col_name):
        idx = col_map.get(col_name)
        if idx is None or idx >= len(row):
            return ''
        v = row[idx]
        return v if v is not None else ''

    def _to_float(v, default=0):
        try:
            if v == '' or v is None:
                return default
            return float(str(v).replace(',', ''))
        except (ValueError, TypeError):
            return default

    def _to_int(v, default=1):
        try:
            if v == '' or v is None:
                return default
            return int(float(str(v).replace(',', '')))
        except (ValueError, TypeError):
            return default

    def _fmt_date(v):
        """엑셀 날짜 → YYYY-MM-DD."""
        if v is None or v == '':
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        s = str(v).strip()
        # '2025.03.01' 또는 '2025/03/01' → '2025-03-01'
        s = s.replace('.', '-').replace('/', '-')
        if len(s) >= 10:
            return s[:10]
        return s

    inserted = 0
    skipped = 0
    comp_set = 0  # 급여항목 설정된 직원 수
    errors = []

    for i, row in enumerate(rows[data_start:], start=data_start + 1):
        if not any(row):
            continue  # 빈 줄
        name = str(_get(row, '이름') or '').strip()
        hire_date = _fmt_date(_get(row, '입사일'))

        if not name:
            errors.append(f'{i}행: 이름 누락')
            continue
        if not hire_date:
            errors.append(f'{i}행({name}): 입사일 누락')
            continue
        # 날짜 형식 검증
        try:
            from datetime import datetime as _dt
            _dt.strptime(hire_date, '%Y-%m-%d')
        except ValueError:
            errors.append(f'{i}행({name}): 입사일 형식 오류 "{hire_date}" (YYYY-MM-DD 필요)')
            continue

        key = (name, hire_date)
        if key in existing_keys:
            skipped += 1
            continue

        is_tax_exempt_raw = str(_get(row, '비과세') or '').strip().upper()
        is_tax_exempt = is_tax_exempt_raw in ('Y', 'YES', 'TRUE', '1', 'O')

        status = str(_get(row, '상태') or '').strip() or '재직'
        if status not in ('재직', '퇴사'):
            status = '재직'

        payload = {
            'name': name,
            'position': str(_get(row, '직급') or '').strip(),
            'department': str(_get(row, '부서') or '').strip(),
            'base_salary': _to_float(_get(row, '기본급'), 0),
            'hire_date': hire_date,
            'status': status,
            'memo': str(_get(row, '메모') or '').strip(),
            'bank_name': str(_get(row, '은행명') or '').strip(),
            'bank_account': str(_get(row, '계좌번호') or '').strip(),
            'dependents_count': _to_int(_get(row, '부양가족수'), 1),
            'is_tax_exempt': is_tax_exempt,
        }

        try:
            emp_result = db.insert_employee(payload)
            inserted += 1
            existing_keys.add(key)

            # 급여 항목 처리
            emp_id = None
            if isinstance(emp_result, dict):
                emp_id = emp_result.get('id')
            elif isinstance(emp_result, list) and emp_result:
                emp_id = emp_result[0].get('id') if isinstance(emp_result[0], dict) else None

            if emp_id:
                components = []
                for kor_label, ctype, is_taxable in _EMP_COMPONENT_COLS:
                    amt = _to_float(_get(row, kor_label), 0)
                    if amt and amt > 0:
                        components.append({
                            'component_type': ctype,
                            'component_name': kor_label,
                            'amount': int(amt),
                            'is_taxable': is_taxable,
                            'is_fixed': True,
                        })
                if components:
                    try:
                        db.bulk_set_salary_components(emp_id, components)
                        comp_set += 1
                    except Exception as ce:
                        errors.append(f'{i}행({name}): 급여항목 저장 실패 - {str(ce)[:80]}')
        except Exception as e:
            errors.append(f'{i}행({name}): 저장 실패 - {str(e)[:80]}')

    _log_action('bulk_upload_employees',
                target=f.filename,
                detail=f'등록 {inserted}건 / 스킵 {skipped}건 / 급여항목 {comp_set}건 / 오류 {len(errors)}건')

    return jsonify({
        'success': True,
        'inserted': inserted,
        'skipped': skipped,
        'components_set': comp_set,
        'errors': errors[:50],  # 최대 50건만
        'error_count': len(errors),
    })


# ══════════════════════════════════════════════
#  급여 관리
# ══════════════════════════════════════════════

@hr_bp.route('/payroll')
@role_required('admin', 'general')
def payroll():
    """급여 관리 메인 페이지"""
    return render_template('hr/payroll.html')


@hr_bp.route('/api/payroll')
@role_required('admin', 'general')
def api_payroll():
    """급여 목록 JSON API"""
    db = get_db()
    pay_month = request.args.get('pay_month', '')
    try:
        rows = db.query_payroll(pay_month=pay_month or None)

        # 직원 이름 매핑 (전체 직원)
        employees = db.query_employees()
        emp_map = {e['id']: e for e in employees}

        # 해당 급여월 기준으로 근무 중이었던 직원만 표시
        # - 재직: 항상 표시
        # - 퇴사: retire_date 가 급여월 이상(해당월 또는 이후)이면 표시
        month_str = pay_month or ''  # 'YYYY-MM' 형태
        enriched = []
        for r in rows:
            eid = r.get('employee_id')
            emp = emp_map.get(eid, {})
            status = emp.get('status', '')
            retire_date = emp.get('retire_date') or ''

            if status == '퇴사' or status == '퇴직':
                # 퇴사일이 급여월보다 이전이면 제외
                if retire_date and month_str and retire_date[:7] < month_str:
                    continue
                # 당월 퇴사 표시
                if retire_date and month_str and retire_date[:7] == month_str:
                    r['emp_status'] = '당월퇴사'
                else:
                    r['emp_status'] = '퇴사'
            elif status not in ('재직', ''):
                continue  # 알 수 없는 상태 제외
            else:
                r['emp_status'] = '재직'

            r['employee_name'] = emp.get('name', '')
            r['department'] = emp.get('department', '')
            r['position'] = emp.get('position', '')
            enriched.append(r)

        total_cost = sum(float(r.get('total_cost', 0)) for r in enriched)

        return jsonify({
            'success': True,
            'payroll': enriched,
            'total_cost': total_cost,
            'count': len(enriched),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/payroll/<int:payroll_id>', methods=['PUT'])
@role_required('admin', 'general')
def api_update_payroll(payroll_id):
    """급여 1건 수정 (수당/메모 수정)"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    try:
        base_salary = float(data.get('base_salary', 0))
        allowances = float(data.get('allowances', 0))
    except (ValueError, TypeError):
        return jsonify({'error': '금액이 올바르지 않습니다.'}), 400

    payload = {
        'base_salary': base_salary,
        'allowances': allowances,
        'total_cost': base_salary + allowances,
        'memo': (data.get('memo') or '').strip(),
    }

    try:
        result = db.update_payroll(payroll_id, payload)
        _log_action('update_payroll', target=f'id={payroll_id}',
                     detail=f'기본급={base_salary:,.0f}, 수당={allowances:,.0f}',
                     new_value=payload)
        return jsonify({'success': True, 'payroll': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/generate', methods=['POST'])
@role_required('admin', 'general')
def api_generate_payroll():
    """월 급여 자동 생성 (한국 급여체계 반영)"""
    db = get_db()
    data = request.get_json() or {}
    pay_month = (data.get('pay_month') or '').strip()
    use_v2 = data.get('use_v2', True)  # 기본적으로 v2 사용

    if not pay_month:
        return jsonify({'error': '대상 월을 지정해주세요.'}), 400

    try:
        if use_v2:
            result = db.generate_monthly_payroll_v2(pay_month)
            inserted = result.get('inserted', 0)
            updated = result.get('updated', 0)
            skipped = result.get('skipped', 0)
            _log_action('generate_payroll',
                         detail=f'대상월={pay_month}, 신규={inserted}건, 갱신={updated}건, 스킵={skipped}건')
            return jsonify({
                'success': True,
                'count': inserted + updated,
                'inserted': inserted,
                'updated': updated,
                'skipped': skipped,
            })
        else:
            count = db.generate_monthly_payroll(pay_month)
            _log_action('generate_payroll',
                         detail=f'대상월={pay_month}, 생성={count}건, v2=False')
            return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/<int:payroll_id>/recalculate', methods=['POST'])
@role_required('admin', 'general')
def api_recalculate_payroll(payroll_id):
    """급여 1건 재계산"""
    db = get_db()
    try:
        result = db.recalculate_payroll(payroll_id)
        if result:
            _log_action('recalculate_payroll', target=f'id={payroll_id}',
                         detail='급여 재계산')
            return jsonify({'success': True, 'payroll': result})
        else:
            return jsonify({'error': '해당 급여를 찾을 수 없습니다.'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/sync-expenses', methods=['POST'])
@role_required('admin', 'general')
def api_sync_expenses():
    """급여 합계를 expenses에 인건비로 자동 반영"""
    db = get_db()
    data = request.get_json() or {}
    pay_month = (data.get('pay_month') or '').strip()

    if not pay_month:
        return jsonify({'error': '대상 월을 지정해주세요.'}), 400

    try:
        result = db.sync_payroll_to_expenses(pay_month)
        _log_action('sync_payroll_to_expenses',
                     detail=f'대상월={pay_month}, 급여={result["total_cost"]:,.0f}, '
                            f'4대보험={result.get("insurance_cost", 0):,.0f}, '
                            f'액션={result["actions"]}')
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/generate-bulk', methods=['POST'])
@role_required('admin', 'general')
def api_generate_bulk_payroll():
    """여러 월 급여 일괄 생성 (입사일~현재 기간)"""
    db = get_db()
    data = request.get_json() or {}
    from_month = (data.get('from_month') or '').strip()
    to_month = (data.get('to_month') or '').strip()

    if not from_month or not to_month:
        return jsonify({'error': '시작월과 종료월을 지정해주세요.'}), 400

    try:
        result = db.generate_bulk_payroll(from_month, to_month)
        _log_action('generate_bulk_payroll',
                     detail=f'{from_month}~{to_month}, '
                            f'신규={result["total_inserted"]}건, 갱신={result["total_updated"]}건')
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  급여 항목 관리 (Salary Components)
# ══════════════════════════════════════════════

@hr_bp.route('/api/employees/<int:emp_id>/salary-components')
@role_required('admin', 'general')
def api_salary_components(emp_id):
    """직원의 급여 항목 목록 조회"""
    db = get_db()
    try:
        components = db.query_salary_components(emp_id, active_only=True)
        return jsonify({'success': True, 'components': components})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/employees/<int:emp_id>/salary-components', methods=['POST'])
@role_required('admin', 'general')
def api_set_salary_components(emp_id):
    """직원의 급여 항목 일괄 설정"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    components = data.get('components', [])

    try:
        count = db.bulk_set_salary_components(emp_id, components)
        _log_action('set_salary_components',
                     target=f'employee_id={emp_id}',
                     detail=f'항목 {count}개 설정')

        # 급여 항목 일괄 변경 → draft 급여 자동 재계산
        recalced = _auto_recalc_payroll(db, emp_id)

        return jsonify({'success': True, 'count': count,
                        'payroll_recalculated': recalced})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/salary-component', methods=['POST'])
@role_required('admin', 'general')
def api_upsert_salary_component():
    """급여 항목 1건 추가/수정"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    employee_id = data.get('employee_id')
    component_type = (data.get('component_type') or '').strip()
    component_name = (data.get('component_name') or '').strip()

    if not employee_id or not component_type:
        return jsonify({'error': '직원과 항목유형은 필수입니다.'}), 400

    try:
        amount = int(float(data.get('amount', 0)))
    except (ValueError, TypeError):
        return jsonify({'error': '금액이 올바르지 않습니다.'}), 400

    payload = {
        'employee_id': int(employee_id),
        'component_type': component_type,
        'component_name': component_name or component_type,
        'amount': amount,
        'is_taxable': data.get('is_taxable', True),
        'is_fixed': data.get('is_fixed', True),
        'effective_from': data.get('effective_from') or date.today().isoformat(),
    }

    comp_id = data.get('id')
    if comp_id:
        payload['id'] = int(comp_id)

    try:
        result = db.upsert_salary_component(payload)
        _log_action('upsert_salary_component',
                     target=f'employee_id={employee_id}',
                     detail=f'{component_name}={amount:,}원',
                     new_value=payload)

        # 급여 항목 변경 → draft 급여 자동 재계산
        recalced = _auto_recalc_payroll(db, employee_id)

        return jsonify({'success': True, 'component': result,
                        'payroll_recalculated': recalced})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/salary-component/<int:comp_id>', methods=['DELETE'])
@role_required('admin', 'general')
def api_delete_salary_component(comp_id):
    """급여 항목 1건 비활성화 (삭제)"""
    db = get_db()
    try:
        # 삭제 전 employee_id 조회 (자동 재계산용)
        emp_id = None
        try:
            res = db.client.table("salary_components").select("employee_id") \
                .eq("id", int(comp_id)).execute()
            if res.data:
                emp_id = res.data[0].get('employee_id')
        except Exception:
            pass

        db.delete_salary_component(comp_id)
        _log_action('delete_salary_component',
                     target=f'id={comp_id}', detail='비활성화')

        # 급여 항목 삭제 → draft 급여 자동 재계산
        recalced = False
        if emp_id:
            recalced = _auto_recalc_payroll(db, emp_id)

        return jsonify({'success': True, 'payroll_recalculated': recalced})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  4대보험 요율 관리 (Insurance Rates)
# ══════════════════════════════════════════════

@hr_bp.route('/api/insurance-rates')
@role_required('admin', 'general')
def api_insurance_rates():
    """4대보험 요율 조회"""
    db = get_db()
    year = request.args.get('year', '')

    try:
        if not year:
            year = date.today().year
        year = int(year)

        rates = db.query_insurance_rates(year=year)
        return jsonify({'success': True, 'rates': rates, 'year': year})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/insurance-rates', methods=['PUT'])
@role_required('admin', 'general')
def api_update_insurance_rates():
    """4대보험 요율 일괄 업데이트"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    year = data.get('year')
    rates = data.get('rates', [])

    if not year or not rates:
        return jsonify({'error': '연도와 요율 데이터는 필수입니다.'}), 400

    try:
        count = db.update_insurance_rates(int(year), rates)
        _log_action('update_insurance_rates',
                     detail=f'연도={year}, 업데이트={count}건')
        return jsonify({'success': True, 'count': count})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/employee-insurance-overrides/<int:employee_id>')
@role_required('admin', 'general')
def api_employee_insurance_overrides(employee_id):
    """직원 개인별 보험요율 오버라이드 조회"""
    db = get_db()
    try:
        overrides = db.query_employee_insurance_overrides(employee_id)
        return jsonify({'success': True, 'overrides': overrides})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/employee-insurance-overrides/<int:employee_id>', methods=['POST'])
@role_required('admin', 'general')
def api_set_employee_insurance_override(employee_id):
    """직원 개인별 보험요율 오버라이드 설정"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    insurance_type = data.get('insurance_type', '')
    employee_rate = data.get('employee_rate')
    employer_rate = data.get('employer_rate')

    if not insurance_type:
        return jsonify({'error': '보험 유형이 필요합니다.'}), 400

    try:
        db.upsert_employee_insurance_override(
            employee_id=employee_id,
            insurance_type=insurance_type,
            employee_rate=float(employee_rate) if employee_rate is not None else 0,
            employer_rate=float(employer_rate) if employer_rate is not None else 0,
            notes=data.get('notes', ''),
        )
        _log_action('set_insurance_override',
                     target=f'employee_id={employee_id}',
                     detail=f'{insurance_type}: 근로자={employee_rate}%, 사업주={employer_rate}%')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/employee-insurance-overrides/<int:employee_id>/delete', methods=['POST'])
@role_required('admin', 'general')
def api_delete_employee_insurance_override(employee_id):
    """직원 개인별 보험요율 오버라이드 삭제 (기본값 복원)"""
    db = get_db()
    data = request.get_json()
    insurance_type = data.get('insurance_type', '') if data else ''

    if not insurance_type:
        return jsonify({'error': '보험 유형이 필요합니다.'}), 400

    try:
        db.delete_employee_insurance_override(employee_id, insurance_type)
        _log_action('delete_insurance_override',
                     target=f'employee_id={employee_id}',
                     detail=f'{insurance_type} 기본값 복원')
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/nontaxable-limits')
@role_required('admin', 'general')
def api_nontaxable_limits():
    """비과세 한도 조회"""
    db = get_db()
    year = request.args.get('year', '')

    try:
        if not year:
            year = date.today().year
        year = int(year)

        limits = db.query_nontaxable_limits(year=year)
        return jsonify({'success': True, 'limits': limits, 'year': year})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/payroll/preview', methods=['POST'])
@role_required('admin', 'general')
def api_payroll_preview():
    """급여 미리보기 (저장 없이 계산 결과만 반환)"""
    db = get_db()
    data = request.get_json() or {}
    employee_id = data.get('employee_id')

    if not employee_id:
        return jsonify({'error': '직원을 지정해주세요.'}), 400

    try:
        from services.hr_service import calculate_payroll

        # 직원 정보
        employees = db.query_employees()
        emp = next((e for e in employees if e['id'] == int(employee_id)), None)
        if not emp:
            return jsonify({'error': '직원을 찾을 수 없습니다.'}), 404

        year = date.today().year
        components = db.query_salary_components(employee_id, active_only=True)
        insurance_rates = db.query_insurance_rates(year=year)
        rate_map = {r['insurance_type']: r for r in insurance_rates}
        nontax_limits = db.query_nontaxable_limits(year=year)
        nontax_map = {r['limit_type']: r['monthly_limit'] for r in nontax_limits}

        # 개인별 보험요율 오버라이드 조회
        overrides = db.query_employee_insurance_overrides(employee_id)
        result = calculate_payroll(emp, components, rate_map, nontax_map,
                                   insurance_overrides=overrides)
        result['employee_name'] = emp.get('name', '')
        result['department'] = emp.get('department', '')
        result['position'] = emp.get('position', '')
        result['insurance_overrides'] = overrides  # 프론트엔드에서 표시용

        return jsonify({'success': True, 'preview': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  급여명세서 PDF
# ══════════════════════════════════════════════

def _enrich_payroll_record(db, record):
    """급여 레코드에 직원 정보 병합 (PDF용)."""
    employees = db.query_employees()
    emp_map = {e['id']: e for e in employees}
    emp = emp_map.get(record.get('employee_id'), {})
    record['employee_name'] = emp.get('name', '')
    record['department'] = emp.get('department', '')
    record['position'] = emp.get('position', '')
    record['hire_date'] = emp.get('hire_date', '')
    return record


@hr_bp.route('/api/payroll/<int:payroll_id>/payslip-pdf')
@role_required('admin', 'general')
def api_payslip_pdf(payroll_id):
    """개별 급여명세서 PDF 다운로드."""
    db = get_db()
    try:
        from reports.payroll_report import generate_individual_payslip

        payroll = db.query_payroll()
        record = next((r for r in payroll if r.get('id') == payroll_id), None)
        if not record:
            return jsonify({'error': '급여 데이터를 찾을 수 없습니다.'}), 404

        record = _enrich_payroll_record(db, record)
        biz_name = current_app.config.get('BIZ_NAME', '배마마')

        fd, path = tempfile.mkstemp(suffix='.pdf')
        os.close(fd)
        generate_individual_payslip(path, record, biz_name=biz_name)

        filename = f"급여명세서_{record.get('employee_name', '')}_{record.get('pay_month', '')}.pdf"
        return send_file(path, as_attachment=True,
                         download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/bulk-payslip-pdf')
@role_required('admin', 'general')
def api_bulk_payslip_pdf():
    """전체 급여명세서 PDF (직원별 페이지)."""
    db = get_db()
    pay_month = request.args.get('pay_month', '')
    if not pay_month:
        return jsonify({'error': '대상 월을 지정해주세요.'}), 400

    try:
        from reports.payroll_report import generate_bulk_payslips

        payroll = db.query_payroll(pay_month=pay_month)
        if not payroll:
            return jsonify({'error': f'{pay_month} 급여 데이터가 없습니다.'}), 404

        employees = db.query_employees()
        emp_map = {e['id']: e for e in employees}
        for r in payroll:
            emp = emp_map.get(r.get('employee_id'), {})
            r['employee_name'] = emp.get('name', '')
            r['department'] = emp.get('department', '')
            r['position'] = emp.get('position', '')
            r['hire_date'] = emp.get('hire_date', '')

        # 이름순 정렬
        payroll.sort(key=lambda r: r.get('employee_name', ''))

        biz_name = current_app.config.get('BIZ_NAME', '배마마')

        fd, path = tempfile.mkstemp(suffix='.pdf')
        os.close(fd)
        generate_bulk_payslips(path, payroll, biz_name=biz_name)

        filename = f"급여명세서_전체_{pay_month}.pdf"
        return send_file(path, as_attachment=True,
                         download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/payroll/summary-pdf')
@role_required('admin', 'general')
def api_payroll_summary_pdf():
    """급여 총괄표 PDF (보고용)."""
    db = get_db()
    pay_month = request.args.get('pay_month', '')
    if not pay_month:
        return jsonify({'error': '대상 월을 지정해주세요.'}), 400

    try:
        from reports.payroll_report import generate_payroll_summary

        payroll = db.query_payroll(pay_month=pay_month)
        if not payroll:
            return jsonify({'error': f'{pay_month} 급여 데이터가 없습니다.'}), 404

        employees = db.query_employees()
        emp_map = {e['id']: e for e in employees}
        for r in payroll:
            emp = emp_map.get(r.get('employee_id'), {})
            r['employee_name'] = emp.get('name', '')
            r['department'] = emp.get('department', '')

        payroll.sort(key=lambda r: r.get('employee_name', ''))
        biz_name = current_app.config.get('BIZ_NAME', '배마마')

        fd, path = tempfile.mkstemp(suffix='.pdf')
        os.close(fd)
        generate_payroll_summary(path, payroll, pay_month, biz_name=biz_name)

        filename = f"급여총괄표_{pay_month}.pdf"
        return send_file(path, as_attachment=True,
                         download_name=filename, mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  퇴직금 계산
# ══════════════════════════════════════════════

@hr_bp.route('/api/employees/<int:emp_id>/severance-calc', methods=['POST'])
@role_required('admin', 'general')
def api_severance_calc(emp_id):
    """퇴직금 계산 (저장 없이 결과만 반환)"""
    db = get_db()
    data = request.get_json() or {}
    retire_date_str = (data.get('retire_date') or '').strip()

    if not retire_date_str:
        retire_date_str = date.today().isoformat()

    try:
        from services.hr_service import calculate_severance

        employees = db.query_employees()
        emp = next((e for e in employees if e['id'] == int(emp_id)), None)
        if not emp:
            return jsonify({'error': '직원을 찾을 수 없습니다.'}), 404

        # 최근 3개월 급여 데이터 조회
        retire_date = date.fromisoformat(retire_date_str)
        recent_payroll = []
        for i in range(1, 4):
            y = retire_date.year
            m = retire_date.month - i
            while m <= 0:
                m += 12
                y -= 1
            pay_month = f'{y}-{m:02d}'
            payroll = db.query_payroll(pay_month=pay_month)
            emp_pay = [p for p in payroll
                       if p.get('employee_id') == emp_id]
            recent_payroll.extend(emp_pay)

        components = db.query_salary_components(emp_id, active_only=True)
        result = calculate_severance(emp, retire_date_str,
                                     recent_payroll, components)
        return jsonify({'success': True, 'severance': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ══════════════════════════════════════════════
#  연차 관리
# ══════════════════════════════════════════════

@hr_bp.route('/leave')
@role_required('admin', 'general')
def leave():
    """연차 관리 메인 페이지"""
    return render_template('hr/leave.html')


@hr_bp.route('/api/leave')
@role_required('admin', 'general')
def api_leave():
    """연차 현황 JSON API (직원별 연차 + 법정일수)"""
    db = get_db()
    year = request.args.get('year', '')

    try:
        if not year:
            from datetime import date
            year = date.today().year

        year = int(year)
        employees = db.query_employees(status='재직')
        all_leave = db.query_annual_leave(year=year)

        # employee_id -> annual_leave 매핑
        leave_map = {r.get('employee_id'): r for r in all_leave}

        result = []
        for emp in employees:
            emp_id = emp['id']
            legal_days = db.calculate_legal_leave_days(emp.get('hire_date'))
            al = leave_map.get(emp_id, {})
            granted = float(al.get('granted_days', 0))
            used = float(al.get('used_days', 0))

            result.append({
                'employee_id': emp_id,
                'employee_name': emp.get('name', ''),
                'department': emp.get('department', ''),
                'hire_date': emp.get('hire_date', ''),
                'legal_days': legal_days,
                'granted_days': granted,
                'used_days': used,
                'remaining_days': granted - used,
            })

        return jsonify({'success': True, 'leave_data': result, 'year': year})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/leave/grant', methods=['POST'])
@role_required('admin', 'general')
def api_grant_leave():
    """연차 부여일수 설정"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    employee_id = data.get('employee_id')
    year = data.get('year')
    granted_days = data.get('granted_days', 0)

    if not employee_id or not year:
        return jsonify({'error': '직원과 연도는 필수입니다.'}), 400

    try:
        granted_days = float(granted_days)
    except (ValueError, TypeError):
        return jsonify({'error': '일수가 올바르지 않습니다.'}), 400

    try:
        result = db.update_annual_leave(employee_id, year, {
            'granted_days': granted_days,
        })
        _log_action('grant_leave',
                     target=f'employee_id={employee_id}',
                     detail=f'연도={year}, 부여={granted_days}일')
        return jsonify({'success': True, 'leave': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/leave', methods=['POST'])
@role_required('admin', 'general')
def api_create_leave():
    """연차 사용 등록"""
    db = get_db()
    data = request.get_json()
    if not data:
        return jsonify({'error': '데이터가 없습니다.'}), 400

    employee_id = data.get('employee_id')
    leave_date = (data.get('leave_date') or '').strip()

    if not employee_id or not leave_date:
        return jsonify({'error': '직원과 날짜는 필수입니다.'}), 400

    try:
        days = float(data.get('days', 1))
    except (ValueError, TypeError):
        days = 1

    payload = {
        'employee_id': int(employee_id),
        'leave_date': leave_date,
        'days': days,
        'leave_type': (data.get('leave_type') or '연차').strip(),
        'memo': (data.get('memo') or '').strip(),
    }

    try:
        result = db.insert_leave_record(payload)
        _log_action('create_leave_record',
                     target=f'employee_id={employee_id}',
                     detail=f'날짜={leave_date}, {days}일',
                     new_value=payload)
        return jsonify({'success': True, 'record': result})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@hr_bp.route('/api/leave/records')
@role_required('admin', 'general')
def api_leave_records():
    """연차 사용 기록 조회"""
    db = get_db()
    employee_id = request.args.get('employee_id', '')
    year = request.args.get('year', '')

    try:
        rows = db.query_leave_records(
            employee_id=int(employee_id) if employee_id else None,
            year=int(year) if year else None,
        )

        # 직원 이름 매핑
        employees = db.query_employees()
        emp_map = {e['id']: e.get('name', '') for e in employees}
        for r in rows:
            r['employee_name'] = emp_map.get(r.get('employee_id'), '')

        return jsonify({'success': True, 'records': rows})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@hr_bp.route('/api/leave/calendar')
@role_required('admin', 'general')
def api_leave_calendar():
    """월별 연차 달력 데이터"""
    db = get_db()
    year = request.args.get('year', '')
    month = request.args.get('month', '')

    if not year or not month:
        from datetime import date
        today = date.today()
        year = year or str(today.year)
        month = month or str(today.month)

    try:
        from services.hr_service import get_leave_calendar
        calendar_data = get_leave_calendar(db, int(year), int(month))
        return jsonify({
            'success': True,
            'calendar': calendar_data,
            'year': int(year),
            'month': int(month),
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
