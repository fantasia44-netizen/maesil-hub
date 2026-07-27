"""입고/생산 엑셀 다운로드 양식 + 업로드 파싱 (미리보기→반영 공용).

재고실사(survey) 패턴과 동일하게:
  - build_*_template : 빈 입력양식(+품목/창고 참조시트) 생성
  - parse_*_excel    : 업로드 파싱 → (items, errors, warnings) 반환 (DB 미반영)
블루프린트의 /excel/preview 가 parse_* 로 미리보기를 만들고,
/excel/apply 가 process_inbound_batch / process_production_batch 로 반영한다.
"""
import io
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ── 스타일 ──
_HDR_FILL = PatternFill('solid', fgColor='4472C4')
_HDR_FONT = Font(color='FFFFFF', bold=True)
_REQ_FILL = PatternFill('solid', fgColor='FFF2CC')   # 필수칸 강조(노랑)
_NOTE_FONT = Font(italic=True, color='999999')

INBOUND_HEADERS = ['날짜(YYYY-MM-DD)', '품목명', '창고위치', '수량',
                   '단위', '카테고리', '보관방법', '비고']
# 필수 컬럼(1-based): 날짜/품목명/창고위치/수량
INBOUND_REQUIRED = {1, 2, 3, 4}

# ★단위/카테고리/보관방법/제조일/소비기한은 '생산품' 속성 — 생산수량 뒤에 묶어 재료와 혼동 방지.
#   재료 블록(재료명/재료수량/재료제조일자)은 끝에 묶음 → 생산품과 안 겹침.
#   생산품 제조일 미입력=생산일자, 소비기한 미입력=제조일+1년 (None 방지).
#   재료제조일자: 미입력=FIFO 차감, 입력=해당 제조일자 배치 우선 차감.
PRODUCTION_HEADERS = ['생산번호', '날짜(YYYY-MM-DD)', '생산위치', '생산품', '생산수량',
                      '단위', '카테고리', '보관방법', '제조일', '소비기한',
                      '재료명', '재료수량', '재료제조일자']
# 필수 컬럼(1-based): 생산번호/날짜/생산위치/생산품/생산수량/재료명/재료수량 (그 외는 선택)
PRODUCTION_REQUIRED = {1, 2, 3, 4, 5, 11, 12}


def _col_letter(idx):
    """1→A, 26→Z, 27→AA …"""
    s = ''
    while idx > 0:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def _style_header(ws, headers, required_idx):
    """required_idx: 필수 컬럼의 1-based 인덱스 집합 → 헤더에 ★ 표시."""
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if c in required_idx:
            cell.value = '★' + h
        ws.column_dimensions[_col_letter(c)].width = 16
    ws.freeze_panes = 'A2'


def _to_date_str(v):
    if v is None or v == '':
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()[:10]


def _to_num(v):
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _s(v):
    return str(v).strip() if v is not None else ''


def _plus_one_year(date_str):
    """제조일 + 1년 (소비기한 기본값). 빈값/오류 시 ''."""
    if not date_str:
        return ''
    try:
        dt = datetime.strptime(str(date_str)[:10], '%Y-%m-%d')
        try:
            return dt.replace(year=dt.year + 1).strftime('%Y-%m-%d')
        except ValueError:        # 2/29 → 2/28
            return dt.replace(year=dt.year + 1, day=28).strftime('%Y-%m-%d')
    except Exception:
        return ''


def _products_for_ref(db):
    """참조시트용 품목 목록 [{name, category, storage_method, unit}]."""
    try:
        prods = db.query_unique_product_names() or []
    except Exception:
        prods = []
    try:
        cost_map = db.query_product_costs() or {}
    except Exception:
        cost_map = {}
    out = []
    seen = set()
    for p in prods:
        nm = p.get('name', '') if isinstance(p, dict) else str(p)
        if not nm or nm in seen:
            continue
        seen.add(nm)
        info = cost_map.get(nm) or cost_map.get(nm.replace(' ', '')) or {}
        out.append({
            'name': nm,
            'category': info.get('category', ''),
            'storage_method': info.get('storage_method', ''),
            'unit': info.get('unit', '개'),
        })
    out.sort(key=lambda x: x['name'])
    return out, set(x['name'] for x in out) | {x['name'].replace(' ', '') for x in out}


def _add_ref_sheets(wb, db, locations):
    products, _ = _products_for_ref(db)
    ws = wb.create_sheet('품목참조')
    ws.append(['품목명', '카테고리', '보관방법', '단위'])
    for c in range(1, 5):
        ws.cell(1, c).font = Font(bold=True)
    for p in products:
        ws.append([p['name'], p['category'], p['storage_method'], p['unit']])
    ws.column_dimensions['A'].width = 26
    for col in 'BCD':
        ws.column_dimensions[col].width = 12

    ws2 = wb.create_sheet('창고참조')
    ws2.append(['창고위치'])
    ws2.cell(1, 1).font = Font(bold=True)
    for loc in (locations or []):
        ws2.append([loc])
    ws2.column_dimensions['A'].width = 16


# ─────────────────────────── 입고 ───────────────────────────

def build_inbound_template(db, locations):
    wb = Workbook()
    ws = wb.active
    ws.title = '입고등록'
    _style_header(ws, INBOUND_HEADERS, INBOUND_REQUIRED)
    _add_ref_sheets(wb, db, locations)

    ws_ex = wb.create_sheet('작성예시')
    ws_ex.append(INBOUND_HEADERS)
    ws_ex.append(['2026-06-18', '동태큐브', '넥스원', 100, '개', '완제품', '냉동', '예시행 — 입고등록 시트에 입력'])
    for c in range(1, len(INBOUND_HEADERS) + 1):
        ws_ex.cell(1, c).font = Font(bold=True)
        ws_ex.cell(2, c).font = _NOTE_FONT
        ws_ex.column_dimensions[_col_letter(c)].width = 16

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def parse_inbound_excel(file, known_names=None):
    """입고 엑셀 → (items, errors, warnings).
    items: [{row, date, product_name, qty, location, unit, category, storage_method, memo}]
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb['입고등록'] if '입고등록' in wb.sheetnames else wb.active
    items, errors, warnings = [], [], []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, '') for v in row):
            continue
        row = list(row) + [''] * (len(INBOUND_HEADERS) - len(row))
        date = _to_date_str(row[0])
        name = _s(row[1])
        location = _s(row[2])
        qty = _to_num(row[3])
        unit = _s(row[4]) or '개'
        category = _s(row[5])
        storage = _s(row[6])
        memo = _s(row[7])

        if not name and not location and qty is None:
            continue
        if not date:
            errors.append(f'{i}행: 날짜가 비었습니다.')
            continue
        if not name:
            errors.append(f'{i}행: 품목명이 비었습니다.')
            continue
        if not location:
            errors.append(f'{i}행: 창고위치가 비었습니다.')
            continue
        if qty is None or qty <= 0:
            errors.append(f'{i}행 ({name}): 수량이 올바르지 않습니다.')
            continue
        if known_names is not None and name.replace(' ', '') not in known_names:
            warnings.append(f'{i}행: \'{name}\'은(는) 상품마스터에 없는 품목입니다(신규 등록됨).')

        items.append({
            'row': i, 'date': date, 'product_name': name, 'qty': qty,
            'location': location, 'unit': unit, 'category': category,
            'storage_method': storage, 'memo': memo,
        })
    return items, errors, warnings


# ─────────────────────────── 생산 ───────────────────────────

def build_production_template(db, locations):
    wb = Workbook()
    ws = wb.active
    ws.title = '생산등록'
    _style_header(ws, PRODUCTION_HEADERS, PRODUCTION_REQUIRED)
    _add_ref_sheets(wb, db, locations)

    ws_ex = wb.create_sheet('작성예시')
    ws_ex.append(PRODUCTION_HEADERS)
    # 생산번호로 그룹: 같은 번호 = 한 생산품 + 재료 여러개 (재료 행마다 반복)
    # 단위/카테고리/보관방법/제조일/소비기한 = 생산품 속성(첫 줄에만), 재료는 명·수량·재료제조일자.
    ex = [
        [1, '2026-06-18', '해서', '동태큐브', 200, '개', '완제품', '냉동', '2026-06-18', '2027-06-17', '동태살', 180, '2026-06-01'],
        [1, '', '', '', '', '', '', '', '', '', '포장재A', 200, ''],
        [2, '2026-06-18', '해서', '대구큐브', 150, '개', '완제품', '냉동', '', '', '대구살', 140, ''],
    ]
    for r in ex:
        ws_ex.append(r)
    for c in range(1, len(PRODUCTION_HEADERS) + 1):
        ws_ex.cell(1, c).font = Font(bold=True)
        ws_ex.column_dimensions[_col_letter(c)].width = 15
    ws_ex.append([])
    ws_ex.append(['※ 같은 생산번호 = 생산품 1건 + 투입재료 여러 줄.'])
    ws_ex.append(['※ 단위·카테고리·보관방법·제조일·소비기한 = \'생산품\' 속성(첫 줄에만). 재료블록(재료명·재료수량·재료제조일자)은 줄마다.'])
    ws_ex.append(['※ 생산품 제조일 미입력 = 생산일자 / 소비기한 미입력 = 제조일+1년 (자동, None 방지).'])
    ws_ex.append(['※ 재료제조일자 미입력 = FIFO 차감 / 입력 시 = 해당 제조일자 배치에서 우선 차감.'])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio


def parse_production_excel(file, known_names=None):
    """생산 엑셀(평면+생산번호 그룹) → (items, errors, warnings).
    items: [{group, date, location, product_name, qty, unit, category,
             storage_method, materials:[{product_name, qty}]}]
    """
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb['생산등록'] if '생산등록' in wb.sheetnames else wb.active
    errors, warnings = [], []
    groups = {}   # gid -> item dict
    order = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or all(v in (None, '') for v in row):
            continue
        row = list(row) + [''] * (len(PRODUCTION_HEADERS) - len(row))
        gid = _s(row[0])
        date = _to_date_str(row[1])
        location = _s(row[2])
        prod = _s(row[3])
        prod_qty = _to_num(row[4])
        unit = _s(row[5]) or '개'        # 생산품 단위
        category = _s(row[6])            # 생산품 카테고리
        storage = _s(row[7])             # 생산품 보관방법
        prod_mfg = _to_date_str(row[8])  # 생산품 제조일 (선택: 미입력=생산일자)
        prod_exp = _to_date_str(row[9])  # 생산품 소비기한 (선택: 미입력=제조일+1년)
        mat = _s(row[10])                # 재료명
        mat_qty = _to_num(row[11])       # 재료수량
        mat_mfg = _to_date_str(row[12])  # 재료 제조일자(선택: 미입력=FIFO, 입력=해당일자 차감)

        if not gid:
            errors.append(f'{i}행: 생산번호가 비었습니다(재료 줄도 생산번호 필요).')
            continue

        g = groups.get(gid)
        if g is None:
            g = {'group': gid, 'date': '', 'location': '', 'product_name': '',
                 'qty': None, 'unit': '개', 'category': '', 'storage_method': '',
                 'manufacture_date': '', 'expiry_date': '',
                 'materials': [], '_rows': []}
            groups[gid] = g
            order.append(gid)
        g['_rows'].append(i)

        # 생산품 정보는 채워진 줄에서 취득(첫 줄)
        if prod and not g['product_name']:
            g['product_name'] = prod
        if prod_qty is not None and g['qty'] is None:
            g['qty'] = prod_qty
        if date and not g['date']:
            g['date'] = date
        if location and not g['location']:
            g['location'] = location
        if unit and g['unit'] == '개':
            g['unit'] = unit
        if category and not g['category']:
            g['category'] = category
        if storage and not g['storage_method']:
            g['storage_method'] = storage
        if prod_mfg and not g['manufacture_date']:
            g['manufacture_date'] = prod_mfg
        if prod_exp and not g['expiry_date']:
            g['expiry_date'] = prod_exp

        # 재료 줄
        if mat:
            if mat_qty is None or mat_qty <= 0:
                errors.append(f'{i}행: 재료 \'{mat}\' 수량이 올바르지 않습니다.')
            else:
                g['materials'].append({'product_name': mat, 'qty': mat_qty,
                                       'manufacture_date': mat_mfg})
                if known_names is not None and mat.replace(' ', '') not in known_names:
                    warnings.append(f'{i}행: 재료 \'{mat}\'은(는) 상품마스터에 없습니다.')

    items = []
    for gid in order:
        g = groups[gid]
        rng = f"(생산번호 {gid})"
        if not g['product_name']:
            errors.append(f'{rng}: 생산품이 비었습니다.')
            continue
        if g['qty'] is None or g['qty'] <= 0:
            errors.append(f'{rng}: 생산수량이 올바르지 않습니다.')
            continue
        if not g['date']:
            errors.append(f'{rng}: 날짜가 비었습니다.')
            continue
        if not g['location']:
            errors.append(f'{rng}: 생산위치가 비었습니다.')
            continue
        if not g['materials']:
            warnings.append(f'{rng} {g["product_name"]}: 투입재료가 없습니다.')
        if known_names is not None and g['product_name'].replace(' ', '') not in known_names:
            warnings.append(f'{rng}: 생산품 \'{g["product_name"]}\'은(는) 상품마스터에 없습니다(신규).')
        # 생산품 제조일/소비기한 기본값 — 빈값 시 None 방지 (수기입력과 동일)
        if not g['manufacture_date']:
            g['manufacture_date'] = g['date']
        if not g['expiry_date']:
            g['expiry_date'] = _plus_one_year(g['manufacture_date'])
        g.pop('_rows', None)
        items.append(g)

    return items, errors, warnings


# ─────────────────────── 정밀 검증 (이름/재고) ───────────────────────

def _normkey(s):
    return str(s or '').replace(' ', '').strip()


def _master_norm_map(db):
    """상품마스터 품목명 {정규화: 원본} 맵."""
    names = []
    try:
        for p in (db.query_unique_product_names() or []):
            nm = p.get('name', '') if isinstance(p, dict) else str(p)
            if nm:
                names.append(nm)
    except Exception:
        pass
    return {_normkey(nm): nm for nm in names}


def _closest(name, norm_map):
    """오타 추천 — 마스터에서 가장 가까운 이름 1개 (없으면 None)."""
    import difflib
    key = _normkey(name)
    if key in norm_map:
        return None
    cand = difflib.get_close_matches(key, list(norm_map.keys()), n=1, cutoff=0.6)
    return norm_map[cand[0]] if cand else None


def _dedup(seq):
    seen, out = set(), []
    for x in seq:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def validate_inbound_items(db, items):
    """입고 항목 이름 검증 — 마스터에 없으면 경고(+오타 추천). 각 item['issues'] 주입."""
    norm = _master_norm_map(db)
    errors, warnings = [], []
    for it in items:
        it.setdefault('issues', [])
        if _normkey(it['product_name']) not in norm:
            sug = _closest(it['product_name'], norm)
            msg = (f"{it['row']}행 '{it['product_name']}': 상품마스터에 없는 품목"
                   + (f" — 혹시 '{sug}'?" if sug else " (신규로 등록됩니다)"))
            it['issues'].append({'level': 'warn', 'field': 'name', 'msg': msg})
            warnings.append(msg)
    return errors, _dedup(warnings)


def validate_production_items(db, items):
    """생산 항목 정밀 검증:
      - 생산품/재료 이름이 마스터에 없으면 경고(+추천)
      - 재료: 위치별 수요 합산 vs 현재고 → 부족/없음이면 오류(반영 차단).
    각 item['issues'] 및 각 material['status']('ok'|'short'|'unknown')·['available'] 주입.
    """
    from services.excel_io import build_stock_snapshot, snapshot_lookup
    norm = _master_norm_map(db)

    snaps = {}

    def _snap(loc):
        if loc not in snaps:
            try:
                snaps[loc] = build_stock_snapshot(db.query_stock_by_location(loc))
            except Exception:
                snaps[loc] = {}
        return snaps[loc]

    # 위치×재료 수요 합산 (배치 전체 — process_production_batch 와 동일 기준)
    demand = {}
    for it in items:
        loc = it.get('location', '')
        for m in it.get('materials', []):
            k = (loc, _normkey(m['product_name']))
            demand[k] = demand.get(k, 0) + (m.get('qty') or 0)

    errors, warnings = [], []
    for it in items:
        loc = it.get('location', '')
        it.setdefault('issues', [])

        # 생산품 이름
        if _normkey(it['product_name']) not in norm:
            sug = _closest(it['product_name'], norm)
            msg = (f"(생산번호 {it['group']}) 생산품 '{it['product_name']}': 마스터에 없음"
                   + (f" — 혹시 '{sug}'?" if sug else " (신규)"))
            it['issues'].append({'level': 'warn', 'field': 'product', 'msg': msg})
            warnings.append(msg)

        # 재료: 이름 + 재고
        for mat in it.get('materials', []):
            mk = _normkey(mat['product_name'])
            sp = snapshot_lookup(_snap(loc), mat['product_name'])
            avail = sp.get('total', 0)
            need = demand.get((loc, mk), 0)
            mat['available'] = avail
            if avail <= 0:
                mat['status'] = 'unknown'
                sug = _closest(mat['product_name'], norm)
                msg = (f"(생산번호 {it['group']}) 재료 '{mat['product_name']}' [{loc}]: "
                       f"마스터/재고에 없음"
                       + (f" — 혹시 '{sug}'?" if sug else ""))
                it['issues'].append({'level': 'error', 'field': 'material', 'msg': msg})
                errors.append(msg)
            elif need > avail:
                mat['status'] = 'short'
                msg = (f"(생산번호 {it['group']}) 재료 '{mat['product_name']}' [{loc}]: "
                       f"재고 부족 — 필요 {need:g} / 재고 {avail:g}")
                it['issues'].append({'level': 'error', 'field': 'material', 'msg': msg})
                errors.append(msg)
            else:
                mat['status'] = 'ok'

    return _dedup(errors), _dedup(warnings)


# ─────────────────────── 반영 그룹핑 (apply 라우팅) ───────────────────────

def group_inbound_for_apply(items, default_date):
    """입고 항목을 날짜별로 그룹핑 (process_inbound_batch 은 단일 날짜)."""
    by_date = {}
    for it in items:
        d = str(it.get('date', '')).strip() or default_date
        by_date.setdefault(d, []).append(it)
    return by_date


def group_production_for_apply(items, default_date):
    """생산 항목을 (날짜, 위치)별로 그룹핑 (process_production_batch 은 단일 날짜+위치)."""
    groups = {}
    for it in items:
        key = (str(it.get('date', '')).strip() or default_date,
               str(it.get('location', '')).strip())
        groups.setdefault(key, []).append(it)
    return groups
